import io
import base64
from pathlib import Path

import pdfplumber
from docx import Document
import openpyxl
from PIL import Image

from app.models.schemas import FileType


def detect_file_type(filename: str, content_type: str) -> FileType:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return FileType.pdf
    elif ext == ".docx":
        return FileType.docx
    elif ext == ".xlsx":
        return FileType.xlsx
    elif ext in (".png", ".jpg", ".jpeg", ".webp"):
        return FileType.image
    # Fallback on content_type
    if "pdf" in content_type:
        return FileType.pdf
    elif "word" in content_type or "document" in content_type:
        return FileType.docx
    elif "sheet" in content_type or "excel" in content_type:
        return FileType.xlsx
    return FileType.image


def extract_pdf_text(file_bytes: bytes) -> str:
    text_parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def extract_docx_text(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_xlsx_text(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))
        if rows:
            text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(text_parts)


def image_to_base64(file_bytes: bytes, content_type: str) -> str:
    # Validate it's a real image
    img = Image.open(io.BytesIO(file_bytes))
    img.verify()
    b64 = base64.b64encode(file_bytes).decode("utf-8")
    return f"data:{content_type};base64,{b64}"


def extract_text(file_bytes: bytes, file_type: FileType) -> str | None:
    try:
        if file_type == FileType.pdf:
            return extract_pdf_text(file_bytes)
        elif file_type == FileType.docx:
            return extract_docx_text(file_bytes)
        elif file_type == FileType.xlsx:
            return extract_xlsx_text(file_bytes)
        return None  # Images don't have extracted text
    except Exception as e:
        return f"[Error extracting text: {e}]"
